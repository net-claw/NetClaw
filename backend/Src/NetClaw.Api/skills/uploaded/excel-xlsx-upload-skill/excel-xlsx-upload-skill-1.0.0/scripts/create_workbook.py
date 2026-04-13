#!/usr/bin/env python3
import argparse
import json
import os
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def autosize_columns(worksheet):
    widths = {}
    for row in worksheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value))

    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 10), 40)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--title", default="NetClaw Export")
    parser.add_argument("--sheet", default="Data")
    parser.add_argument("--json", required=True, dest="json_payload")
    args = parser.parse_args()

    payload = json.loads(args.json_payload)
    headers = payload.get("headers") or ["Name", "Email", "Score"]
    rows = payload.get("rows") or [
        ["Alice", "alice@example.com", 91],
        ["Bob", "bob@example.com", 88],
        ["Carol", "carol@example.com", 95],
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = args.sheet

    worksheet["A1"] = args.title
    worksheet["A1"].font = Font(bold=True, size=14)

    worksheet.append([])
    worksheet.append(headers)
    for cell in worksheet[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")

    for row in rows:
        worksheet.append(row)

    total_row = len(rows) + 4
    if len(headers) >= 3:
        worksheet.cell(row=total_row, column=1).value = "Average score"
        worksheet.cell(row=total_row, column=1).font = Font(bold=True)
        worksheet.cell(row=total_row, column=3).value = f"=AVERAGE(C4:C{total_row - 1})"
        worksheet.cell(row=total_row, column=3).font = Font(bold=True)

    worksheet.freeze_panes = "A4"
    autosize_columns(worksheet)

    meta = workbook.create_sheet("Meta")
    meta["A1"] = "GeneratedAtUtc"
    meta["B1"] = datetime.now(timezone.utc).isoformat()
    meta["A2"] = "Generator"
    meta["B2"] = "NetClaw sandbox example"

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    workbook.save(args.output)


if __name__ == "__main__":
    main()
